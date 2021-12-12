import pandas as pd
import cv2
import openpyxl
import xlwt
import numpy as np
import numpy
import os
import keras
import keras as K
import tensorflow as tf
import matplotlib.pyplot as plt
from keras import optimizers
from keras.models import Model
from keras.optimizers import RMSprop, Adam
from keras.callbacks import LearningRateScheduler
from math import pow, floor
from time import *
from keras.callbacks import ModelCheckpoint
from keras.callbacks import ReduceLROnPlateau
from General.metri1 import SegmentationMetric
from General.metri2 import Precision, Recall, F1
from General.metri3 import conf
from  Network.Unet1 import unet

from General.metri import Dice
from General.metri import dice_similarity
from General.loss_function import soft_dice_loss#soft_dice_loss#focal_loss #focal_loss2
import scipy.io as scio

from sklearn import preprocessing

os.environ["OMP_NUM_THREADS"] = "8"
os.environ["CUDA_VISIBLE_DEVICES"] = '7'  #

def writetoxlsx(data, path):
    # data = open(filename, 'r')
    outwb = openpyxl.Workbook()
    outws = outwb.create_sheet(index=0)
    x = 1
    [h, l] = data.shape
    for i in range(h):
        for j in range(l):
            outws.cell(x, j + 1, data[i, j])
        x += 1
    outwb.save(path)

"""compute learning rate"""


def scheduler(epoch):
    init_lrate = 0.05
    drop = 0.5
    epochs_drop = 10
    lrate = init_lrate * pow(drop, floor(1 + epoch) / epochs_drop)
    print("lr changed to {}".format(lrate))
    return lrate

def save_history(history, result_dir, prefix):
    loss = history.history['loss']
    Dice = history.history['Dice']

    nb_epoch = len(loss)
    # print( nb_epoch)

    with open(os.path.join(result_dir, 'curv_180_test.txt'), 'w') as fp:
        fp.write('epoch\tloss\tDice\n')
        for i in range(nb_epoch):
            fp.write('{}\t{}\t{}\n'.format(
                i, loss[i], Dice[i]))
        fp.close()
        
 


def shapecat(X1, X2):
    a = X1[np.newaxis, :]
    #print(a.shape)
    b = X2[np.newaxis, :]
    c = np.hstack((a, b))
    #print(c[0].shape)
    return c[0]

def NormalizationA(X):
    min_max_scaler=preprocessing.MinMaxScaler()
    X_minMax = min_max_scaler.fit_transform(X)
    return X_minMax

def loadmat(path1,path2,a):
    train_mat1 = path1+str(a)+'.mat'
    data1 = scio.loadmat(train_mat1)
    img1=data1['data']#(165888,17)

    #print(img1.shape)
    train_mat2 = path2+str(a)+'.mat'
    data2 = scio.loadmat(train_mat2)
    img2=data2['data']#(165888,17)
    #print(img2.shape)
    X_i=shapecat(img1, img2)
    #print(X_i)
    img=NormalizationA(X_i)
    #print(img)
    #x=img.reshape((576, 576,16))
    #print(x.shape)
    #img_3d_max=np.amax(img)
    #img_3d_min=np.amin(img)
    #print(img_3d_max)
    #print(img_3d_min)
    #im=(img-img_3d_min)/(img_3d_max-img_3d_min)
    #train_data.append(img)
             
    #print((np.array(train_data)).shape)#(10, 331776, 16)

    
    Train_data=np.array(img)
   
    return Train_data

def load_label(Label1,Label2,a):
    label_mat1 = Label1+str(a)+'.mat'
    data1 = scio.loadmat(label_mat1)
    label_Y1=data1['data']#(1,331776)

    label_mat2 = Label2+str(a)+'.mat'
    data2 = scio.loadmat(label_mat2)
    label_Y2=data2['data']#(1,331776)

    img=shapecat(label_Y1[0], label_Y2[0])
        
    #print(label_Y[0].shape)#(331776,)
    #Y=label_Y[0].reshape((576, 576))
    #train_label.append(img)
    
    #print((np.array(train_label)).shape)#(10, 331776)

    
    Train_label=np.array(img)
    return Train_label


def tran_data(data):
    (x, y,z) = data.shape
    CoorA = np.zeros(((x * y),z), dtype=np.float64)
    #print(CoorA.shape)  # (663552,)
    for i in range(x):
        CoorA[(i * y):((i + 1) * y)] = data[i]

    #print('CoorA.shape=',CoorA.shape)
    return CoorA

def sample_data(data, sample_size,size):
    #print('data.shape',data.shape)
    (x, y) = data.shape  # (20, 331776,16)
    #print('data.shape=',data.shape)
    #Data = tran_data(data)  # (6635520,16)
    a = int((x / sample_size))  # 648
    #print('a=',a)#3240
    coorA = np.zeros((a, sample_size,y), dtype=np.float64)
    #print(coorA.shape)  # (648, 1024)
    
    for j in range(a):
        coorA[j] = data[(j * sample_size):((j + 1) * sample_size)]
        #print(coorA[j])
    #print('coorA.shape=',coorA.shape)
    coor=coorA.reshape(a,size,size,y)
    #print('coor.shape=',coor.shape)
    return coor


def tran_label(data):
    (x, y) = data.shape
    CoorA = np.zeros(((x * y)), dtype=np.float64)
    #print(CoorA.shape)  # (663552,)
    for i in range(x):
        CoorA[(i * y):((i + 1) * y)] = data[i]

    #print(CoorA.shape)
    return CoorA


def sample_label(data, sample_size,size):
    #print('data.shape',data.shape)
    y = (data.shape)[0]  # (2, 331776)
    #print('y',y)
    #Data = tran_label(data)  # (663552,)
    a = int((y / sample_size) )  # 648
    #print(a)#648
    coorA = np.zeros((a, sample_size), dtype=np.float64)
    #print(coorA.shape)  # (648, 1024)
    for j in range(a):
        coorA[j] =data[(j * sample_size):((j + 1) * sample_size)]
        #print(coorA[j])
    #print(coorA.shape)
    coor=coorA.reshape(a,size,size)
    return coor
 
"""
Input=(var_len,sample_size,2) #(162, 4096, 2)
output=(20,331776,2)
"""
def tran_shape(Coor):
    # print(Coor)
    (x, y, z) = Coor.shape
    CoorA = np.zeros(((x * y), z), dtype=np.float64)
    # print(CoorA.shape) #(331776, 3)
    for i in range(x):
        CoorA[(i * y):((i + 1) * y)] = Coor[i]
    # print(CoorA)
    return CoorA


def predict_upsample(predict, num):
    #print("predict.shape=",predict.shape) ##(81, 4096, 2)
    (a, b, c) = predict.shape #202, 128, 128, 2
    predict_a = tran_shape(predict)  # (81*4096, 2)=(331776,2)
    """
    d = int((a * b))
    #print(d)
    predictA = np.zeros((var_xlen, d, c), dtype=np.float64)  # (2,331776,2)
    #print(predictA.shape)  # (20,331776,2)

    for i in range(var_xlen):
        #print(predict_a[i * d:(i + 1) * d])
        #print(predictA[i])
        predictA[i] = predict_a[i * d:(i + 1) * d]
    """
    return predict_a




####multi_feature
def load_mat(path,a,batch):
    b=a*batch
    img=[]
    for j in range(batch):
        c=b+j
        #print('c=',c)
        train_mat1 = path+str(c)+'.mat'
        data1 = scio.loadmat(train_mat1)
        img1=data1['data']#(165888,17)
        img.append(img1)
    Train_data=np.array(img)
    #print('Train_data.shape',Train_data.shape)
    return Train_data
def data_generator(path1,label,number,batchsize):
    while(True):
         for i in range(number):
             
             DataA=load_mat(path1,i,batchsize)
             #print('DataA.shape',DataA.shape)
             
             
             Label=load_mat(label,i,batchsize)

             X1 = DataA[:,:, :,:]
             
             Y = Label[:, :,:]
             y = keras.utils.to_categorical(Y, num_classes=2)
             yield X1,y


def var_data_generator(var_path1,var_label,var_number,var_batchsize):
    while(True):
         for j in range(var_number):
             var_DataA=load_mat(var_path1,j,var_batchsize)
             #print('DataA.shape',DataA.shape)
             
             
             var_Label=load_mat(var_label,j,var_batchsize)

             var_X1 = var_DataA[:,:, :,:]
            
             var_Y = var_Label[:, :,:]
             var_y = keras.utils.to_categorical(var_Y, num_classes=2)
             yield var_X1,var_y

if __name__ == '__main__':
    #path_xls='/storage/student22/net_hinge/mymodel/xlsx/sulc_volume_16_std_U_SE_24.xlsx'
    #filepath = "/storage/student22/net_hinge/mymodel/sulc_volume_16_std_U_SE_24.h5"
    #filepath2 = "/storage/student22/net_hinge/mymodel/sulc_volume_16_std_save_U_SE_24.h5"#_e
    #path_xls='/storage/student22/net_hinge/mymodel/xlsx/sulc_volume_16_std_U_SE_24.xlsx'
    filepath = "/storage/student22/net_hinge/mymodel/sulc_16_std_U_SE_24_seg.h5"
    filepath2 = "/storage/student22/net_hinge/mymodel/sulc_16_std_save_U_SE_24_seg.h5"#_e


    #filepath = "/storage/ZLL/code/mymodel/curv_16_std_U_SE_24_seg.h5"#_ex
    #filepath2 = "/storage/ZLL/code/mymodel/thickness_sulc_16_std_save_U_SE_24_seg.h5"#_e
    #path_xls='/storage/ZLL/code/mymodel/xlsx/2add_16_std_U_SE_24_t.xlsx'
    #####  data_load ####a
    
    # T1
    train_1='/storage/student22/3hinge_seg/1000/new_data/16/seg/train/sulc/'
    test_1='/storage/student22/3hinge_seg/1000/new_data/16/seg/test/sulc/'

    #train_2='/storage/student22/3hinge_seg/1000/new_data/16/seg/train/area/'
    #test_2='/storage/student22/3hinge_seg/1000/new_data/16/seg/test/area/'
    #label
    train_label='/storage/student22/3hinge_seg/1000/new_data/16/seg/label/'
    test_label='/storage/student22/3hinge_seg/1000/new_data/16/seg/test_label/'

  
    vect=16
    #para
    xlen =1822
    var_xlen =202
    num=var_xlen
    size=64
    Batch_size=40
    batch=1
    #epoch_a=int((331776*xlen)/sample_size)
    #epoch_l=int((331776)/sample_size)
    #print('epoch_l=',epoch_l)
    epoch=10
    #####load_data####
    Data=data_generator(train_1,train_label,xlen,Batch_size)

    var_Data=var_data_generator(test_1,test_label,var_xlen,Batch_size)

    #print(Data.__next__())
    #X1,X2,Y=Data.next()
    #print(Data('X1'))
       
    ##### model  ####
    T1_shape = (size,size,vect)#(sample_size,16)
    #T2_shape = (size,size,vect)#(sample_size,16)
    model =unet(T1_shape) #concat_net(T1_shape,coor_shape) #unet(T1_shape)
    model.summary()
    RMSprop = optimizers.RMSprop(lr=0.05, rho=0.9, epsilon=1e-4, decay=0.0)  ###lr=0.05 importance
    
    #Adad=optimizers.Adadelta(lr=1.0, rho=0.95, epsilon=1e-06)
    model.compile(optimizer='RMSprop', loss=soft_dice_loss, metrics=[Dice])#categorical_crossentropy binary_crossentropy
    learningrate = LearningRateScheduler(scheduler)
    checkpoint = ModelCheckpoint(filepath, monitor='val_loss', verbose=1, save_best_only=True,save_weights_only=True, mode='auto')
    callbacks_list = [checkpoint, learningrate]
    #history=model.fit([sulc_train_data,curv_train_data],train_label,validation_data=([sulc_test_data,curv_test_data],test_label),batch_size=Batch_size,epochs=epoch,callbacks=callbacks_list,shuffle=True)

    history=model.fit_generator(Data,steps_per_epoch=xlen//batch,epochs=epoch,callbacks=callbacks_list,
            validation_data=var_Data,validation_steps=var_xlen//batch)
    model.save(filepath2)



    plt.plot(history.history['Dice'])
    plt.plot(history.history['val_Dice'])
    plt.title('Model Dice')
    plt.ylabel('Dice')
    plt.xlabel('Epoch')
    plt.legend(['Train', 'Val'], loc='upper left')
    plt.show()

    plt.plot(history.history['loss'])
    plt.plot(history.history['val_loss'])
    plt.title('Model loss')
    plt.ylabel('Loss')
    plt.xlabel('Epoch')
    plt.legend(['Train', 'Val'], loc='upper left')
    plt.show()
