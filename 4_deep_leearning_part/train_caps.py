import pandas as pd
import cv2
import openpyxl
import xlwt
import numpy as np
import numpy
import os
import keras
import keras as K
import tensorflow as tf
import matplotlib.pyplot as plt
from keras import optimizers
from keras.models import Model
from keras.optimizers import RMSprop, Adam
from keras.callbacks import LearningRateScheduler
from math import pow, floor
from time import *
from keras.callbacks import ModelCheckpoint
from keras.callbacks import ReduceLROnPlateau
from General.metri1 import SegmentationMetric
from General.metri2 import Precision, Recall, F1
from General.metri3 import conf
from  Network.new_capsnet import capsnet#unet
#from  Network.U_caps import unet
#from  Network.unet_models import U_CapsNet
from General.metri import Dice
from General.metri import dice_similarity
from General.loss_function import soft_dice_loss#soft_dice_loss#focal_loss #focal_loss2
import scipy.io as scio

from sklearn import preprocessing


os.environ["CUDA_VISIBLE_DEVICES"] = '3'  #

def writetoxlsx(data, path):
    # data = open(filename, 'r')
    outwb = openpyxl.Workbook()
    outws = outwb.create_sheet(index=0)
    x = 1
    [h, l] = data.shape
    for i in range(h):
        for j in range(l):
            outws.cell(x, j + 1, data[i, j])
        x += 1
    outwb.save(path)

"""compute learning rate"""


def scheduler(epoch):
    init_lrate = 0.05
    drop = 0.5
    epochs_drop = 10
    lrate = init_lrate * pow(drop, floor(1 + epoch) / epochs_drop)
    print("lr changed to {}".format(lrate))
    return lrate

def save_history(history, result_dir, prefix):
    loss = history.history['loss']
    Dice = history.history['Dice']

    nb_epoch = len(loss)
    # print( nb_epoch)

    with open(os.path.join(result_dir, 'curv_180_test.txt'), 'w') as fp:
        fp.write('epoch\tloss\tDice\n')
        for i in range(nb_epoch):
            fp.write('{}\t{}\t{}\n'.format(
                i, loss[i], Dice[i]))
        fp.close()
        
 


def shapecat(X1, X2):
    a = X1[np.newaxis, :]
    #print(a.shape)
    b = X2[np.newaxis, :]
    c = np.hstack((a, b))
    #print(c[0].shape)
    return c[0]

def NormalizationA(X):
    min_max_scaler=preprocessing.MinMaxScaler()
    X_minMax = min_max_scaler.fit_transform(X)
    return X_minMax

def loadmat(path1,path2,num):
    train_data = []  
    for i in range(num):
             train_mat1 = path1+str(i)+'.mat'
             data1 = scio.loadmat(train_mat1)
             img1=data1['data']#(165888,17)



             #print(img1.shape)
             train_mat2 = path2+str(i)+'.mat'
             data2 = scio.loadmat(train_mat2)
             img2=data2['data']#(165888,17)
             #print(img2.shape)
             X_i=shapecat(img1, img2)
             #print(X_i)
             img=NormalizationA(X_i)
             #print(img)
             #x=img.reshape((576, 576,16))
             #print(x.shape)
             #img_3d_max=np.amax(img)
             #img_3d_min=np.amin(img)
             #print(img_3d_max)
             #print(img_3d_min)
             #im=(img-img_3d_min)/(img_3d_max-img_3d_min)
             train_data.append(img)
             
    #print((np.array(train_data)).shape)#(10, 331776, 16)

    
    Train_data=np.array(train_data)
    Train_data=Train_data[:, :,:]
    return Train_data

def load_label(Label1,Label2,num):
    train_label = []
    for i in range(num):
        label_mat1 = Label1+str(i)+'.mat'
        data1 = scio.loadmat(label_mat1)
        label_Y1=data1['data']#(1,331776)

        label_mat2 = Label2+str(i)+'.mat'
        data2 = scio.loadmat(label_mat2)
        label_Y2=data2['data']#(1,331776)

        img=shapecat(label_Y1[0], label_Y2[0])
        
        #print(label_Y[0].shape)#(331776,)
        #Y=label_Y[0].reshape((576, 576))
        train_label.append(img)
    
    print((np.array(train_label)).shape)#(10, 331776)

    
    Train_label=np.array(train_label)
    return Train_label


def tran_data(data):
    (x, y,z) = data.shape
    CoorA = np.zeros(((x * y),z), dtype=np.float64)
    #print(CoorA.shape)  # (663552,)
    for i in range(x):
        CoorA[(i * y):((i + 1) * y)] = data[i]

    #print('CoorA.shape=',CoorA.shape)
    return CoorA

def sample_data(data, sample_size,size):
    (num,x, y) = data.shape  # (20, 331776,16)
    #print('data.shape=',data.shape)
    Data = tran_data(data)  # (6635520,16)
    a = int((x / sample_size) * num)  # 648
    #print('a=',a)#3240
    coorA = np.zeros((a, sample_size,y), dtype=np.float64)
    #print(coorA.shape)  # (648, 1024)
    
    for j in range(a):
        coorA[j] = Data[(j * sample_size):((j + 1) * sample_size)]
        #print(coorA[j])
    #print('coorA.shape=',coorA.shape)
    coor=coorA.reshape(a,size,size,y)
    #print('coor.shape=',coor.shape)
    return coor


def tran_label(data):
    (x, y) = data.shape
    CoorA = np.zeros(((x * y)), dtype=np.float64)
    #print(CoorA.shape)  # (663552,)
    for i in range(x):
        CoorA[(i * y):((i + 1) * y)] = data[i]

    #print(CoorA.shape)
    return CoorA


def sample_label(data, sample_size,size):
    (x, y) = data.shape  # (2, 331776)
    Data = tran_label(data)  # (663552,)
    a = int((y / sample_size) * x)  # 648
    # print(a)#648
    coorA = np.zeros((a, sample_size), dtype=np.float64)
    #print(coorA.shape)  # (648, 1024)
    for j in range(a):
        coorA[j] = Data[(j * sample_size):((j + 1) * sample_size)]
        #print(coorA[j])
    #print(coorA.shape)
    coor=coorA.reshape(a,size,size)
    return coor
 
"""
Input=(var_len,sample_size,2) #(162, 4096, 2)
output=(20,331776,2)
"""
def tran_shape(Coor):
    # print(Coor)
    (x, y, z) = Coor.shape
    CoorA = np.zeros(((x * y), z), dtype=np.float64)
    # print(CoorA.shape) #(331776, 3)
    for i in range(x):
        CoorA[(i * y):((i + 1) * y)] = Coor[i]
    # print(CoorA)
    return CoorA


def predict_upsample(predict, var_xlen):
    #print(predict.shape)
    (a, b, c) = predict.shape #202, 128, 128, 2
    predict_a = tran_shape(predict)  # (162*4096,16,2)
    d = int((a * b) / var_xlen)
    #print(d)
    predictA = np.zeros((var_xlen, d, c), dtype=np.float64)  # (2,331776,2)
    #print(predictA.shape)  # (20,331776,2)

    for i in range(var_xlen):
        #print(predict_a[i * d:(i + 1) * d])
        #print(predictA[i])
        predictA[i] = predict_a[i * d:(i + 1) * d]
    
    return predictA
def margin_loss(y_true, y_pred):
    lamb, margin = 0.5, 0.1
    result = K.backend.sum(y_true * K.backend.square(K.backend.relu(1 - margin -y_pred))
    + lamb * (1-y_true) * K.backend.square(K.backend.relu(y_pred - margin)), axis=-1)
    return result
          
if __name__ == '__main__':
    path_xls='/storage/student22/net_hinge/mymodel/xlsx/sulc_16_std_Caps.xlsx'
    filepath = "/storage/student22/net_hinge/mymodel/sulc_16_std_Caps.h5"
    filepath2 = "/storage/student22/net_hinge/mymodel/sulc_16_std_save_Caps.h5"#_e
    
    filepath = "/storage/ZLL/code/mymodel/sulc_16_std_Caps.h5"#_ex
    filepath2 = "/storage/ZLL/code/mymodel/sulc_16_std_Caps.h5"#_e
    #path_xls='/storage/ZLL/code/mymodel/xlsx/sulc_16_std_Caps.xlsx'
    #####  data_load ####a
    
    # T1
    train_l ='/storage/student22/3hinge_seg/1000/new_data/16/train/sulc/lh/'
    train_r ='/storage/student22/3hinge_seg/1000/new_data/16/train/sulc/rh/'

    test_l ='/storage/student22/3hinge_seg/1000/new_data/16/test/sulc/lh/'
    test_r ='/storage/student22/3hinge_seg/1000/new_data/16/test/sulc/rh/'


    label_l ='/storage/student22/3hinge_seg/1000/new_data/label/lh/'
    label_r ='/storage/student22/3hinge_seg/1000/new_data/label/rh/'
    label_test_l ='/storage/student22/3hinge_seg/1000/new_data/label_test/lh/'
    label_test_r ='/storage/student22/3hinge_seg/1000/new_data/label_test/rh/'
    vect=16
    #para
    xlen =10
    var_xlen =2
    num=var_xlen
    batch_size=1
    sample_size=1024*4
    size=64
    Batch_size=40
    epoch_a=int((331776*xlen)/sample_size)
    epoch_l=int((331776*var_xlen)/sample_size)
    print('epoch_a=',epoch_a)
    epoch=5
    #print('epoch_l=',epoch_l)

    #####load_data####



    #test
       #data
    test_dataA=loadmat(test_l,test_r,var_xlen)
    #print('train_dataA=',test_dataA.shape) ##(10, 331776, 16)
    test_data=sample_data(test_dataA, sample_size,size) #
    #print('test_data=',test_data.shape)#((810, 64, 64, 16))
       #label
    test_labelA=load_label(label_test_l,label_test_r,var_xlen)  ##(10, 331776, 16)
    #print('test_labelA=',test_labelA.shape)
    test_labelB=sample_label(test_labelA, sample_size,size)#(810, 64, 64, 16)
    test_label= keras.utils.to_categorical(test_labelB, num_classes=2)

    #train
       #data
    train_dataA=loadmat(train_l,train_r,xlen)
    #print('train_dataA=',train_dataA.shape) ##(20, 331776, 16)
    train_data=sample_data(train_dataA, sample_size,size) #
    #print('train_data=',train_data.shape)#((1620, 64, 64, 16))
       #label
    train_labelA=load_label(label_l,label_r,xlen)  ##(20, 331776)
    train_labelB=sample_label(train_labelA, sample_size,size)#(1620, 64, 64)
    train_label= keras.utils.to_categorical(train_labelB, num_classes=2)#(1620, 64, 64,2)
    
    ##### model  ####
    T1_shape = (size,size,vect)#(sample_size,16)
    #model=unet(T1_shape)
    model=capsnet(T1_shape, 2, 3)
    #ucaps =U_CapsNet(T1_shape, n_class = 2, decoder = False)#CapsNet(T1_shape) #concat_net(T1_shape,coor_shape) #unet(T1_shape)
    #model =ucaps.build_model(model_layer = "simple", capsnet_type = "r3", upsamp_type = "subpix")
    model.summary()
    RMSprop = optimizers.RMSprop(lr=0.05, rho=0.9, epsilon=1e-4, decay=0.0)  ###lr=0.05 importance
    
    #Adad=optimizers.Adadelta(lr=1.0, rho=0.95, epsilon=1e-06)
    model.compile(optimizer='RMSprop', loss=soft_dice_loss, metrics=[Dice])#categorical_crossentropy binary_crossentropy#soft_dice_loss
    learningrate = LearningRateScheduler(scheduler)
    checkpoint = ModelCheckpoint(filepath, monitor='val_loss', verbose=1, save_best_only=True,save_weights_only=True, mode='auto')
    callbacks_list = [checkpoint, learningrate]
    
    history=model.fit(train_data,train_label,validation_data=(test_data,test_label),
            batch_size=Batch_size,epochs=epoch,callbacks=callbacks_list,shuffle=True)
    model.save(filepath2)
    ### evaluate ###
    
    predictionsA = model.predict(test_data)
    #print('predictionsA.shape=',predictionsA.shape)#(202, 128, 128, 2)
    predictionsB=predictionsA.reshape(epoch_l,sample_size,2)
    #print('predictionsB.shape=',predictionsB.shape)
    predictions=predict_upsample(predictionsB, var_xlen)
    #print('predictions.shape=',predictions.shape)
    BG1 = []
    hinge_List1 = []
    prec1 = []
    reca1 = []
    F_11 = []
    Fpr1 = []
    Fnr1 = []
    Tp1 = []
    Fp1 = []
    Fn1 = []
    Tn1 = []
    acc1= []
    IoU1= []
    mIoU1= []
    traget=[]
    for i in range(num):
        print(predictions[i])
        print(predictions[i].shape)
        segmentationImage2A = np.argmax(predictions[i], axis=1)
        print(segmentationImage2A)  #segmentationImage2A.shape=(576,576)
        print(segmentationImage2A.shape)
        dice_hinge1, dice_BG1 = dice_similarity(test_labelA[i], segmentationImage2A)
        print('dice_BG1', dice_BG1)
        print('dice_hinge1', dice_hinge1)
        (recall1, precision1, F11, FPR1, FNP1, TP1, FP1, FN1, TN1,acc,IoU,mIoU) = conf(test_labelA[i], segmentationImage2A)

        print('precision', precision1)
        print('recall', recall1)
        print('F1', F11)

        print('FPR', FPR1)
        print('FNP', FNP1)
        print('TP:', TP1)
        print('FP:', FP1)
        print('FN:', FN1)
        print('TN:', TN1)
        print('acc:', acc)
        print('IoU:', IoU)
        print('mIoU:', mIoU)
        
        traget.append(segmentationImage2A)
        BG1.append(dice_BG1)
        hinge_List1.append(dice_hinge1)
        prec1.append(precision1)
        reca1.append(recall1)
        F_11.append(F11)
        Fpr1.append(FPR1)
        Fnr1.append(FNP1)
        Tp1.append(TP1)
        Fp1.append(FP1)
        Fn1.append(FN1)
        Tn1.append(TN1)
        acc1.append(acc)
        IoU1.append(IoU)
        mIoU1.append(mIoU)
    print("%s: %.4f (+/- %.4f)" % ("BG", numpy.mean(BG1), numpy.std(BG1)))
    print("%s: %.4f (+/- %.4f)" % ("hinge", numpy.mean(hinge_List1), numpy.std(hinge_List1)))
    print("%s: %.4f (+/- %.4f)" % ("precision", numpy.mean(prec1), numpy.std(prec1)))
    print("%s: %.4f (+/- %.4f)" % ("recall", numpy.mean(reca1), numpy.std(reca1)))
    print("%s: %.4f (+/- %.4f)" % ("F1", numpy.mean(F_11), numpy.std(F_11)))
    print("%s: %.4f (+/- %.4f)" % ("FPR", numpy.mean(Fpr1), numpy.std(Fpr1)))
    print("%s: %.4f (+/- %.4f)" % ("FNR", numpy.mean(Fnr1), numpy.std(Fnr1)))
    print("%s: %.2f (+/- %.4f)" % ("TP", numpy.mean(Tp1), numpy.std(Tp1)))
    print("%s: %.2f (+/- %.4f)" % ("FP", numpy.mean(Fp1), numpy.std(Fp1)))
    print("%s: %.2f (+/- %.4f)" % ("FN", numpy.mean(Fn1), numpy.std(Fn1)))
    print("%s: %.2f (+/- %.4f)" % ("TN", numpy.mean(Tn1), numpy.std(Tn1)))
    print("%s: %.4f (+/- %.4f)" % ("acc1", numpy.mean(acc1), numpy.std(acc1)))
    print("%s: %.4f (+/- %.4f)" % ("IoU1", numpy.mean(IoU1), numpy.std(IoU1)))
    print("%s: %.4f (+/- %.4f)" % ("mIoU1", numpy.mean(mIoU1), numpy.std(mIoU1)))

    print(filepath)
    
    target3=np.array(traget)
    print(target3.shape)
    #target=re_shape(target3)
    #print(target.shape)
    target_1=target3.T
    #writetoxlsx(target_1, path_xls)
    

    plt.plot(history.history['Dice'])
    plt.plot(history.history['val_Dice'])
    plt.title('Model Dice')
    plt.ylabel('Dice')
    plt.xlabel('Epoch')
    plt.legend(['Train', 'Val'], loc='upper left')
    plt.show()

    plt.plot(history.history['loss'])
    plt.plot(history.history['val_loss'])
    plt.title('Model loss')
    plt.ylabel('Loss')
    plt.xlabel('Epoch')
    plt.legend(['Train', 'Val'], loc='upper left')
    plt.show()

  
