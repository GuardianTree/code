import pandas as pd
import numpy as np
import os
import xlwt
import openpyxl
from time import *

os.environ["OMP_NUM_THREADS"] = "5"  #use CPU
 


def writetoxlsx(data, path):
    #data = open(filename, 'r')
    outwb = openpyxl.Workbook()
    outws = outwb.create_sheet(index=0)
    x = 1
    [h, l]=data.shape
    for i in range(h):
        for j in range(l):
            outws.cell(x, j + 1,data[i,j])
        x += 1
    outwb.save(path)  


def load_coor(path,n):
    hinge=pd.read_excel(path)
    print('coor load OK!!')
    lenver=len(hinge["ver"+str(0)].values) 
    a=0
    x1=np.zeros((n, lenver), dtype=np.float64)
    b=1
    y1=np.zeros((n, lenver), dtype=np.float64)
    c=2
    z1=np.zeros((n, lenver), dtype=np.float64)  
    for i in range(n):
        x1[i] = hinge["ver"+str(a)].values 
        y1[i] = hinge["ver"+str(b)].values 
        z1[i] = hinge["ver"+str(c)].values
        
        a=a+3
        b=b+3
        c=c+3   
    
    return (np.array(x1)),(np.array(y1)),(np.array(z1))
def load_coorA(path,n):
    hinge=pd.read_excel(path)
    print('coor load OK!!')
    lenver=len(hinge["ver"+str(0)].values) 
    a=0
    x1=np.zeros((n, lenver), dtype=np.float64)
    b=1
    y1=np.zeros((n, lenver), dtype=np.float64)
    c=2
    z1=np.zeros((n, lenver), dtype=np.float64)
    v=0
    a1=np.zeros((n, lenver), dtype=np.int64)  
    for i in range(n):
        x1[i] = hinge["ver"+str(a)].values 
        y1[i] = hinge["ver"+str(b)].values 
        z1[i] = hinge["ver"+str(c)].values
        a1[i] = hinge["log"+str(v)].values
        a=a+3
        b=b+3
        c=c+3   
        v=v+1
    return (np.array(a1)),(np.array(x1)),(np.array(y1)),(np.array(z1))
def load_coorB(path,n):
    hinge=pd.read_excel(path)
    print('coor load OK!!')
    lenver=len(hinge["log"+str(0)].values) 
    v=0
    a1=np.zeros((n, lenver), dtype=np.int64)  
    for i in range(n):
        
        a1[i] = hinge["log"+str(v)].values
          
        v=v+1
    return (np.array(a1))
def get_orig(x,y,z,n):

    for a in range(n):
        lenx=len(x[a]) 

    A=np.zeros((n,lenx,3), dtype=np.float64)
    
    for i in range(n):
        #print(len(x[i]))
        for j in range(len(x[i])):
            A[i][j][0]=(x[i][j])
            A[i][j][1]=(y[i][j])
            A[i][j][2]=(z[i][j])
   
    return A

def EDistanceA(A,C,n,a2,a3):
    for a in range(n):
        xlen_A=len(A[a])
    
    for a in range(n):
        xlen_C=len(C[a])
    arr=np.array([0,0,0]).astype(float)

    c=[]
    c2=[]
    mesh=[]
    for i in range(n):
        
        mesh1=[]
        mesh2=[]
        for j in range(xlen_C):
            
            begin_time1=time()
            coords=C[i][j]
            np_c = np.array(coords)
            print(np_c)
            if ((np_c).tolist()!=arr.tolist()):
                #print(arr)
                print(i)
                print(j)
                print('ok')
                for k in range(xlen_A):#xlen_A
                    coords1=A[i][k]
                    np_c1 = np.array(coords1)
                
                    d1=np.sqrt(np.sum((np_c - np_c1 )**2))
                    #print(d1)
                    if d1<=2.0:
                       mesh.append(d1)
                       mesh1.append(k) 
                    elif d1<=8.0:
                         a=k+1
             
                         if a in a2[i]:
                            #print(a)
                            #print(a2[i])
                            mesh2.append(k)
            else:
                print('Not OK!!!')
            
            end_time1=time()
            run_time1 =end_time1-begin_time1
            
            print('the for  cost time is :',run_time1)

            print(len(mesh1))
            print(len(mesh2))
        c.append(mesh1)
        c2.append(mesh2)
        
    ca=np.array(c)
    cb=np.array(c2)
    return ca,cb
def dataTo(data,a,b):
    x=a
    y=b
    dataq=np.zeros((x,y))
    for i in range(x):
        for j in range(len(data[i])):
            dataq[i,j]=data[i][j]
        for d in range(len(data[i]),y):
            dataq[i,d]=data[i][0]
    return dataq
    
if __name__ == '__main__':
        path_coor='/storage/ZLL/code/label_test/coor/3h_seg/ex_data/0_99/data/coor_l_0_99.xlsx'
        path_3hinge='/storage/ZLL/code/label_test/coor/3h_seg/ex_data/0_99/data/3h_l_0_99.xlsx'
        path_2hinge='/storage/ZLL/code/label_test/coor/3h_seg/ex_data/0_99/data/2h_l_0_99.xlsx'

        path_target='/storage/ZLL/code/label_test/coor/3h_seg/ex_data/0_99/3hinge_l_0_99.xlsx'
        path_target1='/storage/ZLL/code/label_test/coor/3h_seg/ex_data/0_99/2hinge_l_0_99.xlsx'
        #coor data load
        begin_time = time()
        n=100
        b=20000
        (x0,y0,z0)=load_coor(path_coor,n)
        A=get_orig(x0,y0,z0,n)
        #a2=load_coorB(path_2hinge,n)
        (a2,x2,y2,z2)=load_coorA(path_2hinge,n)
        B=get_orig(x2,y2,z2,n)
        (a3,x3,y3,z3)=load_coorA(path_3hinge,n)
        C=get_orig(x3,y3,z3,n)
        
        print(A)
        print(A.shape)
        #print(B)
        #print(B.shape)
        print(C)
        print(C.shape)
        print(a2) 
        print(a3)
        (mesh,meshB)=EDistanceA(A,C,n,a2,a3)
        
        dataA=dataTo(mesh,n,b)
        mesh0=dataA.T
        print(mesh0.shape)
        writetoxlsx(mesh0,path_target)

        dataB=dataTo(meshB,n,b)
        meshb=dataB.T
        print(meshb.shape)
        writetoxlsx(meshb,path_target1)
        end_time = time()
        run_time =end_time-begin_time
        print('the code cost time is :',run_time)










