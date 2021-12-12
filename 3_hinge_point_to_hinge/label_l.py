import pandas as pd
import numpy as np
import os
import xlwt
import openpyxl
from time import * 
#os.environ["OMP_NUM_THREADS"] = "5"  #use CPU
#path_label='/storage/student22/label_an/label/label_r_test.xlsx'

path_2h0='/storage/ZLL/code/label_test/coor/3h_seg/ex_data/0_99/2hinge_l_0_99.xlsx'
path_2h1='/storage/ZLL/code/label_test/coor/3h_seg/ex_data/0_99/3hinge_l_0_99.xlsx'
path_2h2='/storage/ZLL/code/label_test/coor/3h_seg/ex_data/0_99/mesh_2hinge_l.xlsx'
path_2h3='/storage/ZLL/code/label_test/coor/3h_seg/ex_data/0_99/data/3h_l_0_99.xlsx'

target_label='/storage/ZLL/code/label_test/coor/3h_seg/ex_data/0_99/label_l_0_99.xlsx'
def writetoxlsx(data, path):
    #data = open(filename, 'r')
    outwb = openpyxl.Workbook()
    outws = outwb.create_sheet(index=0)
    x = 1
    [h, l]=data.shape
    for i in range(h):
        for j in range(l):
            outws.cell(x, j + 1,data[i,j])
        x += 1
    outwb.save(path)


#hinge data load
df_2hinge0=pd.read_excel(path_2h0)
print('2hinge0_ok!')
#hinge data load
df_2hinge1=pd.read_excel(path_2h1)
print('2hinge1_ok!')
df_2hinge2=pd.read_excel(path_2h2)
print('2hinge2_ok!')
df_2hinge3=pd.read_excel(path_2h3)
print('2hinge3_ok!')

target3=[]
target=[]
n=165888
num=100
for i in range(num):
    begin_time = time()
    
    head_data='data'+str(i)
    head_data1='log'+str(i)
    hinge01_2 = df_2hinge0[head_data].values
    hinge0_2  = list(map(int,hinge01_2))
    hinge11_2 = df_2hinge1[head_data].values
    hinge1_2  = list(map(int,hinge11_2))
    hinge21_2 = df_2hinge2[head_data].values
    hinge2_2  = list(map(int,hinge21_2))
    hinge31_2 = df_2hinge3[head_data1].values
    hinge3_2  = list(map(int,hinge31_2))
    print(hinge01_2.shape)
    print(i)
    
    target2=[]
    for j in range(n):
        a=j+1
        if j in hinge0_2:
            da=1
        elif j in hinge1_2:
            da=1
        elif j in hinge2_2:
            da=1
        elif j in hinge3_2:
            da=1
        else:
            da=0
        target2.append(da)
    target3.append(target2)
    end_time = time()
    run_time =end_time-begin_time
    print('the code cost time is :',run_time)
target3=np.array(target3)
print(target3)
print(target3.shape)
target=target3.T
print(target.shape)
#data_write(path_target, target)
writetoxlsx(target,target_label)

   
