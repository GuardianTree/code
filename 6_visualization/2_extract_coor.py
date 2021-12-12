import pandas as pd
import numpy as np
import numpy
import os
import openpyxl
def load_data(path,xlen):
    #data load
    df1 = pd.read_excel(path)
    xline_l=len(df1["data"+str(0)].values)
    print("data1ok?")
    x= np.zeros((xlen,xline_l), dtype=np.int64)
    for i in range(xlen):
        x[i] = df1["data"+str(i)].values
    print("ok!")
    return x


def load_coor(path,n):
    hinge=pd.read_excel(path)
    print('coor load OK!!')
    lenver=len(hinge["ver"+str(0)].values) 
    a=0
    x1=np.zeros((n, lenver), dtype=np.float64)
    b=1
    y1=np.zeros((n, lenver), dtype=np.float64)
    c=2
    z1=np.zeros((n, lenver), dtype=np.float64)
    v=0
    v1=np.zeros((n, lenver), dtype=np.int64)  
    for i in range(n):
        x1[i] = hinge["ver"+str(a)].values 
        y1[i] = hinge["ver"+str(b)].values 
        z1[i] = hinge["ver"+str(c)].values
        v1[i] = hinge["log"+str(v)].values
        a=a+3
        b=b+3
        c=c+3
        v=v+1   
    return v1,x1,y1,z1
def extract_b(index,log,x,y,z,n):
    for a in range(n):
        len_index=len(index[a])
        len_log=len(log[a])
        print(len_index)
        print(len_log)
    label_x=[]
    label_y=[]
    label_z=[]
    for i in range(n):
        relog=[]
        Log=[]
        X=[]
        Y=[]
        Z=[]
        print(i)
        for j in range(len_index):
            #print(index[i][j])
            
            a=int(index[i][j])
            if a not in relog:
                b=(log[i].tolist()).index(a)
                X.append(x[i][b])
                #print(X)
                Y.append(y[i][b])
                #print(Y)
                Z.append(z[i][b])
                #print(Z)
                Log.append(b)
                #print(Log)
                #print(b)
            relog.append(a)
        label_x.append(X)
        label_y.append(Y)   
        label_z.append(Z)    
    return label_x,label_y,label_z
def extract_a(index,label,a,b):
    (x,y)=label.shape
    Label_a=[]
    Feature_a=[]
    for i in range(x):
        Label=[]
      
        for j in range(y):
            if j in index[i]:
               Label.append(label[i,j])
              
        print(len(Label))
        for k in range(len(Label),b):
            if len(Label)<b:
                Label.append(0)
        print(len(Label))
        Label_a.append(Label)


    label_a=[]
    label_a=np.array(Label_a)
    

    return label_a
            
def writetoxlsx(data, path,n,kk):
    #data = open(filename, 'r')
    outwb = openpyxl.Workbook()
    outws = outwb.create_sheet(index=0)
    x = 1
    
    print(data.shape)
    dataA=data.T
    print(dataA.shape)
    for i in range(kk):
        
        #l = len(data[i])
         
        for j in range(n):
            outws.cell(x, j + 1,dataA[i][j])
        x += 1
    outwb.save(path)  


def dataTo(data,a,b):
    dataq=np.zeros((a,b))
    for i in range(a):
        y=len(data[i].tolist())
        for j in range(y):
            dataq[i,j]=data[i][j]
    return dataq
def shapecat(X1,X2,xlen):
    x=xlen
    print(x)
    a = []
    b = []
    c = []
    d=[] 
    for i in range(x):
        a = X1[i][np.newaxis,:]
        b = X2[i][np.newaxis,:]
        c=np.hstack((a,b))
        d.append(c)
    d= np.asarray(d)
    print(d.shape)
    return d
def get_orig(x,y,z,n,xline):

    #for a in range(n):
    #    lenx=len(x[a]) 

    A=np.zeros((n*3,xline), dtype=np.float64)
    a=0
    b=1
    c=2
    for i in range(n):
        #print(len(x[i]))
        
        for j in range(len(x[i])):
            A[a][j]=(x[i][j])
            A[b][j]=(y[i][j])
            A[c][j]=(z[i][j])

        a=a+3
        b=b+3
        c=c+3
        
    
    return A
"""
def get_orig(x,y,z,n):

    for a in range(n):
        lenx=len(x[a]) 

    A=np.zeros((n,lenx,3), dtype=np.float64)
    
    for i in range(n):
        #print(len(x[i]))
        for j in range(len(x[i])):
            A[i][j][0]=(x[i][j])
            A[i][j][1]=(y[i][j])
            A[i][j][2]=(z[i][j])
   
    return A
"""
if __name__ == '__main__':
        path_index='/storage/student22/code/see/sulc_seg_2048_U1_index.xlsx'
        path_coor='/storage/student22/code/see/coor_l_testB.xlsx'
        target_coor='/storage/student22/code/see/my_l_coor_03.xlsx'
        xlen=1

        df_label=pd.read_excel(path_index)
        log=df_label['data0'].values
        print(len(log))
        xline=len(log)

        index=load_data(path_index,xlen)
        print(index)
        print((np.array(index)).shape)
        (log,x,y,z)=load_coor(path_coor,xlen)
        print(log)
        print(log.shape)
        print(x)

        
        (Label_x,Label_y,Label_z)=extract_b(index,log,x,y,z,xlen)
        #print(len(Label_x[1]))
        B=get_orig(Label_x,Label_y,Label_z,xlen,xline)
        
        dataLabel=dataTo(B,xlen*3,xline)
       
        writetoxlsx(dataLabel,target_coor,xlen*3,xline)
        

        
